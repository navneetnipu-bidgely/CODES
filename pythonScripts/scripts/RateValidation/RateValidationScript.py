#!/bin/python3

"""
    # inputs:
        # The proper formatted rate sheet (unique format for any pilot)
    # output:
        # generates an Excel sheet in the same format as of input Excel file
        # with data indicating the gaps in rate time bands  for particular rate band id and valid time-band

    # approach:
        # Read the required columns of the Excel sheet provided as input.
        # Iterate over each sheet one by one available in the parent Excel sheet
        # Each sheet will be operated by one thread
        # Then iterate over each row of the sheet and fetch each planNumber's different rate bands and their data
        # Then give the complete rate plans data fetched to validate_rate_plan function
        # The validate_rate_plan function will call separate functions to validate the gaps in year or month or week or day
        # Each of these functions will operate in multithreading
        # Finally the validate_rate_plan will get the gaps in rate plans and call the write_data_to_excel_sheet function
          to write the data into Excel sheet that will be the outputof this program.
        # Also, we can generatea plot visualizing the gaps in a particular rate plan.

"""
import datetime
# import the below required packages necessary to run this progarm.

# packages required to handle json data structure or put api responses into json
import json

# packages required for executing multithreading
import threading

# packages required to handle Excel sheet write operations
from xlwt import Workbook

# packages required to handle Excel sheet read operations
import openpyxl

# package required to handle arrays
import array

# import the user inputs from file UserInputs.py
import UserInputs

# packages required to handle logging operations like create log file, write in log file at different log level like info,debug,warning etc
import logging

# import RateDataStructure to use rate plan proper data structure to store respective
# months,days,hours,week days for respective rate band
import RateDataStructure

RATE_PLAN_DATA_INITIALIZATION=RateDataStructure.RATE_PLAN_DATA_STRUCTURE
MONTH_NUMBER_NAME_MAP=RateDataStructure.MONTH_NUMBER_NAME_MAP
WEEK_NUMBER_NAME_MAP=RateDataStructure.WEEK_DAYS_NUMBER_NAME_MAP

# output variable

"""
    RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]={
        "missing_months":[],
        "missing_week_days":{month:[week days]},...},
        "missing_days":{month:[],...},
        "missing_hours":{month:{day:[]},...}
    }
"""

RATE_PLAN_VALIDATION_OUTPUT_DATA={}

# color printing color constant
RED_COLOR="\033[1;31;40m"
GREEN_COLOR="\033[1;32;40m"
BLUE_COLOR="\033[1;34;40m"
DEFAULT_COLOR="\033[0m"
YELLOW_BACKGROUND_GREEN_COLOR="\033[1;32;43m"


# create a log file
DIRECTORY=UserInputs.LOG_FILE_PATH

# get the current date and time
current_datetime = datetime.datetime.now()

# format the current date and time as a string
formatted_datetime = current_datetime.strftime("%d%m%Y%H%M%S")

LOG_FILE_PATH= DIRECTORY + "RatePlanValidationScriptReport"+ formatted_datetime+".log"

# creating the required log file
with open(LOG_FILE_PATH, 'w') as file:
    pass


# Configure the logging module
# change the level config for different log level like log,debug,warning etc.
logging.basicConfig(filename=LOG_FILE_PATH, level=logging.INFO)


# function to read Excel sheet data
def read_excelSheet_data(rate_sheet_file_path):
    """

    :param rate_sheet_file_path:
    :return:
    """

    # fetches each sheet data and return them as a list data structure
    # creating an empty list for storing all sheets data
    excel_sheet_data_as_list=[]

    # open and read Excel sheet from given rate_sheet_file_path path
    try:
        workbook=openpyxl.load_workbook(rate_sheet_file_path)
    except Exception as e:
        print(RED_COLOR," Exception found while loading the excel sheet from given path:",rate_sheet_file_path)
        print(e,DEFAULT_COLOR)

    # getting all the sheet names in the above given sheet
    try:
        sheet_names=workbook.sheetnames

        print("following sheets has to be validated for Rates:\n",sheet_names)

    except Exception as e:
        print(RED_COLOR," Exception found while getting the sheets names in the given sheet:",rate_sheet_file_path)
        print(e,DEFAULT_COLOR)

    # iterating over each sheet present in the given Excel sheet
    for sheetName in sheet_names:

        try:

            # active sheet
            active_sheet=workbook[sheetName]

            """
            # forming column header and column number mapping so as to access
            # cell value using column header names

            column_headers_with_column_number_mapping={}

            # getting cell value from 1st row of active_sheet
            for cell in active_sheet[1]:
                column_headers_with_column_number_mapping[cell.value]=cell.column
            """

            # Now storing all active_sheet data into a list
            # Create an empty list
            active_sheet_data = []

            # Loop through the rows and columns of the sheet
            for row in active_sheet.iter_rows(values_only=True):
                row_data = []
                for cell in row:
                    # using column number (UserInputs.RATE_COLUMN_NUMBER) 17 as it is rate or charge
                    # which should be in float
                    # should be changed as per sheet column header sequence for rate
                    if row.index(cell)!=UserInputs.RATE_COLUMN_NUMBER-1 and isinstance(cell, float):
                        row_data.append(int(cell))
                    else:
                        row_data.append(cell)
                active_sheet_data.append(row_data)

            # appending active_sheet_data to excel_sheet_data_as_list
            excel_sheet_data_as_list.append(active_sheet_data)


        except Exception as e:
            print(RED_COLOR," Exception occured while getting data from sheet:", sheetName)
            print(e,DEFAULT_COLOR)


    # Close the workbook
    workbook.close()

    return excel_sheet_data_as_list


# todo function to write data in excel sheet
def write_data_to_excel_sheet():
    pass

# function to validate rate plans
def validate_rate_plan(*sheet_data):
    """

    :param sheet_data:
    :return:
    """


    # this function will accept sheet_data as list and convert it to tuple of lists
    # where each list will be a single row data.
    # now we will iterate over each row and fixed cell numbers to scan the whole sheet data
    # needed to validate the rate plan.

    rate_plan_data=generate_rate_validator_data_from_sheet_data(sheet_data=sheet_data)
    logging.info(json.dumps(rate_plan_data, indent=4))
    # print(YELLOW_BACKGROUND_GREEN_COLOR,rate_plan_data,DEFAULT_COLOR)

    # proceed with the validation of rate_plan_data

    validate_month_gaps(rate_data=rate_plan_data)
    validate_week_day_gaps(rate_data=rate_plan_data)
    validate_day_gaps(rate_data=rate_plan_data)
    validate_hour_gaps(rate_data=rate_plan_data)

    print(BLUE_COLOR,"Validation done for following unique rate bands:",rate_plan_data.keys(),DEFAULT_COLOR)
    # print(GREEN_COLOR,"Rate plan validation done for sheet data:\n",sheet_data,DEFAULT_COLOR)



# function to scan the complete sheet and store the required values for validation
def generate_rate_validator_data_from_sheet_data(sheet_data):

    """
    :param sheet_data:
    :return: rate_plan_data
    """

    rate_plan_data = {}

    try:

        # forming column header to column number mapping to access cell values using column headers
        column_headers_to_column_number_mapping = {}
        for cell in sheet_data[0]:
            column_headers_to_column_number_mapping[cell] = sheet_data[0].index(cell)

        # operating on sheet_data but skipping 1st row that has been used for column headers
        # to generate column_headers_to_column_number_mapping.
        for row_data in sheet_data[1:]:
            # each row data will be in form of list

            unique_key=find_unique_key_for_rate_band(row_data,column_headers_to_column_number_mapping)

            # if unique key not present in the rate_band_id, then initialize it
            if unique_key not in rate_plan_data:
                rate_plan_data[unique_key] = RATE_PLAN_DATA_INITIALIZATION

            # now populating in rate_plan_data the respective applicable month numbers for respective rate band key
            month_start = row_data[column_headers_to_column_number_mapping["monthLow"]]
            month_end = row_data[column_headers_to_column_number_mapping["monthHigh"]]

            for month_number in range(month_start, month_end + 1):
                rate_plan_data[unique_key][MONTH_NUMBER_NAME_MAP[month_number]]['present']=True

            # similarly populating in rate_plan_data the respective applicable day numbers for respective rate band key
            day_start = row_data[column_headers_to_column_number_mapping["dayLow"]]
            day_end = row_data[column_headers_to_column_number_mapping["dayHigh"]]

            for month_number in range(month_start, month_end + 1):
                for day_number in range(day_start, day_end + 1):
                    rate_plan_data[unique_key][MONTH_NUMBER_NAME_MAP[month_number]]['days'][day_number]['present'] = True

            # similarly populating in rate_plan_data the respective applicable week day numbers for respective rate band key
            week_day_start = row_data[column_headers_to_column_number_mapping["weekLow"]]
            week_day_end = row_data[column_headers_to_column_number_mapping["weekHigh"]]

            for month_number in range(month_start, month_end + 1):
                for week_day_number in range(week_day_start, week_day_end + 1):
                    rate_plan_data[unique_key][MONTH_NUMBER_NAME_MAP[month_number]]['week_days'][WEEK_NUMBER_NAME_MAP[week_day_number]] = True


            # similarly populating in rate_plan_data the respective applicable hours for respective rate band key
            hour_start = row_data[column_headers_to_column_number_mapping["timeOfDayLow"]]
            hour_end = row_data[column_headers_to_column_number_mapping["timeOfDayHigh"]]

            for month_number in range(month_start, month_end + 1):
                for day_number in range(day_start, day_end + 1):
                    for hour_number in range(hour_start, hour_end+1):
                        rate_plan_data[unique_key][MONTH_NUMBER_NAME_MAP[month_number]]['days'][day_number]['hours'][hour_number] = True


    except Exception as e:
        print(RED_COLOR," Exception occured while scanning sheet with data:",DEFAULT_COLOR)
        print(sheet_data)
        print(RED_COLOR,e,DEFAULT_COLOR)

    # return the rate_plan_data to parent function for validation
    return rate_plan_data

# function to find the unique_key for a rate band id
def find_unique_key_for_rate_band(row_data,column_headers):
    # we need to identify the fields which make a rate band id unique
    # for each row we will be having different planNumber,rateBandId,validLow,validHigh
    # initializing them before reading the cell data

    plan_number = row_data[column_headers["planNumber"]]
    rate_band_id = row_data[column_headers["rateBandId"]]
    valid_start_date = str(row_data[column_headers["validLow"]])
    valid_end_date = str(row_data[column_headers["validHigh"]])
    # consumption_low=str(row_data[column_headers["consumptionLow"]])
    # consumption_high=str(row_data[column_headers["consumptionHigh"]])
    is_holiday=str(row_data[column_headers["isHoliday"]])
    # rate=str(row_data[column_headers["rate"]])
    # threshold=str(row_data[column_headers["threshold"]])
    # tou_name=str(row_data[column_headers["touName"]])
    tier_name = str(row_data[column_headers["tierName"]])
    # group_name=str(row_data[column_headers["groupName"]])

    unique_key="planNumber:"+str(plan_number) + '|' + "rateBandId:"+str(rate_band_id) + '|' + "validLow:"+valid_start_date + '|' + "validHigh:"+valid_end_date + '|'+"is_holiday:"+str(is_holiday)+'|' + "tier:"+tier_name

    return unique_key


# todo function to validate month gaps
def validate_month_gaps(rate_data):
    """
    :param rate_data:
    :return:
    """

    try:
        # for each unique rate band, check the month field
        # if present key of each month is true that means that this month is not included in that rate band

        for unique_key in rate_data:
            # initialize the RATE_PLAN_VALIDATION_DATA for unique_key is not present
            if unique_key not in RATE_PLAN_VALIDATION_OUTPUT_DATA:
                RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]={}
                RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_months"]=[]

            months=rate_data[unique_key]

            for month in months:
                # for every month, if months[month]['present']==False then add the month in missing month list
                if not months[month]['present']:
                    RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_months"].append(month)

        print(GREEN_COLOR,"month wise validation done for following unique rate bands:\n",rate_data.keys(),DEFAULT_COLOR)

    except Exception as e:
        print(RED_COLOR," Exception occured while validating month gaps for rate band :\n", unique_key)
        print(e,DEFAULT_COLOR)

# todo function to validate week gaps
def validate_week_day_gaps(rate_data):
    try:
        # for each unique rate band, check the week field of each month field
        # if week_day key for week is true that means that this week is included in that rate band for this month

        for unique_key in rate_data:
            # RATE_PLAN_VALIDATION_DATA is already initialized in validate_month_gaps for each available unique key.
            # initialize the RATE_PLAN_VALIDATION_DATA of a unique_key for missing_week_days is not present
            if "missing_week_days" not in RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]:
                RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_week_days"] = {}

            months = rate_data[unique_key]


            for month in months:
                # for every month, if we need to iterate over 'week_days' key to find if a week_day is present or not
                # for a particular month.
                week_days=months[month]['week_days']
                for day_of_week in week_days:
                    if not week_days[day_of_week]:
                        if month not in RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_week_days"]:
                            RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_week_days"][month]=[]
                        RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_week_days"][month].append(day_of_week)

        print(GREEN_COLOR, "week wise validation done for following unique rate bands:\n", rate_data.keys(), DEFAULT_COLOR)

    except Exception as e:
        print(RED_COLOR," Exception occured while validating week gaps for rate band :\n", unique_key)
        print(e,DEFAULT_COLOR)

# todo function to validate day gaps
def validate_day_gaps(rate_data):
    """
    :param rate_data:
    :return:
    """

    try:
        # for each unique rate band, check the days field of each month field
        # if present key for a day of a month is true that means that this day is included in that rate band for this month

        for unique_key in rate_data:
            # RATE_PLAN_VALIDATION_DATA is already initialized in validate_month_gaps for each available unique key.
            # initialize the RATE_PLAN_VALIDATION_DATA of a unique_key for missing_days if not present
            if "missing_days" not in RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]:
                RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_days"] = {}

            months = rate_data[unique_key]

            for month in months:
                # for every month, if we need to iterate over 'days' key to find if a day is present or not
                # for a particular month.
                days = months[month]['days']
                for day in days:
                    if not days[day]['present']:

                        if month not in RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_days"]:
                            RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_week_days"][month] = []

                        RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_days"][month].append(day)

        print(GREEN_COLOR, "day wise validation done for following unique rate bands:\n", rate_data.keys(), DEFAULT_COLOR)
    except Exception as e:
        print(RED_COLOR," Exception occured while validating day gaps for rate band :\n", unique_key)
        print(e,DEFAULT_COLOR)

# todo function to validate hour gaps
def validate_hour_gaps(rate_data):
    """
    :param rate_data:
    :return:
    """

    try:
        # for each unique rate band, check the days field of each month field
        # if present key for a day of a month is true that means that this day is included in that rate band for this month

        for unique_key in rate_data:
            # RATE_PLAN_VALIDATION_DATA is already initialized in validate_month_gaps for each available unique key.
            # initialize the RATE_PLAN_VALIDATION_DATA of a unique_key for missing_hours if not present
            if "missing_hours" not in RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]:
                RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_hours"] ={}

            months = rate_data[unique_key]

            for month in months:
                # for every month, if we need to iterate over 'days' key to find and then over each day's
                # hours key to find the missing hours if thatbparticular day is included in the rate band
                # for a particular month.
                days = months[month]['days']
                for day in days:
                    if days[day]['present']:
                        hours=days[day]['hours']
                        for hour in hours:
                            if not hours[hour]:
                                print(hour)
                                if month not in RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_hours"]:
                                    RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_week_days"][month] = {}
                                if day not in RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_hours"][month]:
                                    RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_week_days"][month][day] =[]

                                RATE_PLAN_VALIDATION_OUTPUT_DATA[unique_key]["missing_days"][month][day].append(hour)

        print(GREEN_COLOR, "hour wise validation done for following unique rate bands:\n", rate_data.keys(), DEFAULT_COLOR)

    except Exception as e:
        print(RED_COLOR," Exception occured while validating hour gaps for rate band :\n", unique_key)
        print(e,DEFAULT_COLOR)

# generate and execute threads
def thread_generator_and_executor(all_sheet_data,thread_limit):
    """

    :param all_sheet_data:
    :param thread_limit:
    :return:
    """

    # generates thread for each sheet and stores the threads in thread_collection
    # once thread_collection size reaches the thread_limit, they will be executed first and them emptied
    # again repeat the same process untill we reach the last sheet

    threads_collection = []
    number_of_sheets=len(all_sheet_data)

    try:
        # now iterating over each sheet in all_sheet_data

        for sheet in all_sheet_data:
            # create a thread with the thread target function for each sheet
            thread_object = threading.Thread(target=validate_rate_plan, args=sheet)
            threads_collection.append(thread_object)

            # when threads_collection size reach thread_limit or sheet_number reach the last sheet
            # all the threads in threads_collection will be executed by thread_executor() function
            # and them emptied

            if len(threads_collection) == thread_limit or all_sheet_data.index(sheet)+1 == (number_of_sheets):
                # execute the threads
                for thread in threads_collection:
                    thread.start()
                    thread.join()

                print("Threads executed for sheet number upto", all_sheet_data.index(sheet)+1)

                # todo store the outputs from the thread_executor

                # making the threads_collection empty when all threads in threads_collection are being executed
                threads_collection = []

        print("All sheets have been processed by threads successfully")

    except Exception as e:
        print(RED_COLOR," Exception occured while processing sheet:", all_sheet_data.index(sheet))
        print(e,DEFAULT_COLOR)


if __name__=='__main__':

    print("Into the main function")

    print("Calling read_excelSheet_data() function to fetch excel sheet data as list")
    sheet_data=read_excelSheet_data(rate_sheet_file_path=UserInputs.INPUT_FILE_PATH)

    print("Calling thread_generator_and_executor function to generate threads and execute them")

    thread_generator_and_executor(all_sheet_data=sheet_data,thread_limit=UserInputs.MAX_THREAD_LIMIT)

    print(BLUE_COLOR,"final rate validation data:")
    print(GREEN_COLOR,json.dumps(RATE_PLAN_VALIDATION_OUTPUT_DATA, indent=4),DEFAULT_COLOR)

    logging.info("RATE_PLAN_VALIDATION_OUTPUT_DATA:\n")
    logging.info(RATE_PLAN_VALIDATION_OUTPUT_DATA)

    print(RED_COLOR,"main thread function executed.")






